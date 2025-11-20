#!/usr/bin/env python3
"""
GÉNÉRATEUR ANNEXE 2 AUTOMATIQUE - GIRASOLE 2025
Diagnostic Photovoltaïque

Génère l'ANNEXE 2 (fichier de synthèse) automatiquement depuis les checklists terrain
Format: Excel conforme à l'exemple fourni par Girasole
"""

import json
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

def charger_data_audit_json(json_terrain_path, json_be_path):
    """
    Charge les données d'audit depuis les JSON terrain + BE
    
    Args:
        json_terrain_path: Chemin vers JSON checklist terrain
        json_be_path: Chemin vers JSON checklist BE
    
    Returns:
        Dict avec données audit consolidées
    """
    data_audit = {
        'statut': 'À compléter',
        'anomalies_critiques': 0,
        'anomalies_majeures': 0,
        'anomalies_mineures': 0,
        'nb_photos': 0,
        'conformite_iec': 'À vérifier',
        'recommandations': 'À compléter',
        'observations': 'À compléter'
    }
    
    # Charger terrain JSON si existe
    if json_terrain_path.exists():
        try:
            with open(json_terrain_path, 'r', encoding='utf-8') as f:
                terrain = json.load(f)
                data_audit['nb_photos'] = terrain.get('nb_photos_prises', 0)
                data_audit['observations'] = terrain.get('observations_terrain', 'À compléter')
        except Exception as e:
            print(f"   ⚠️  Erreur lecture {json_terrain_path.name}: {e}")
    
    # Charger BE JSON si existe (prioritaire sur terrain)
    if json_be_path.exists():
        try:
            with open(json_be_path, 'r', encoding='utf-8') as f:
                be = json.load(f)
                data_audit['statut'] = be.get('statut_audit', 'À compléter')
                data_audit['anomalies_critiques'] = be.get('nb_anomalies_critiques', 0)
                data_audit['anomalies_majeures'] = be.get('nb_anomalies_majeures', 0)
                data_audit['anomalies_mineures'] = be.get('nb_anomalies_mineures', 0)
                data_audit['conformite_iec'] = be.get('conformite_IEC_62446', 'À vérifier')
                data_audit['recommandations'] = be.get('recommandations_prioritaires', 'À compléter')
                # BE observations prioritaires si présentes
                if be.get('observations_be'):
                    data_audit['observations'] = be.get('observations_be')
        except Exception as e:
            print(f"   ⚠️  Erreur lecture {json_be_path.name}: {e}")
    
    return data_audit

def creer_annexe2_structure():
    """Crée structure Excel ANNEXE 2 conforme exemple Girasole"""
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Synthèse Audits"
    
    # Styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # En-têtes colonnes
    headers = [
        "ID Centrale",
        "Nom Projet",
        "Puissance (kWc)",
        "Type",
        "Département",
        "Date Audit",
        "Auditeur",
        "Statut Audit",
        "Anomalies Critiques",
        "Anomalies Majeures",
        "Anomalies Mineures",
        "Photos Prises",
        "Conformité IEC 62446-3",
        "Recommandations Prioritaires",
        "Observations Terrain"
    ]
    
    # Écrire en-têtes
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(1, col_idx, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    # Ajuster largeurs colonnes
    col_widths = [12, 30, 15, 20, 12, 12, 20, 15, 18, 18, 18, 12, 20, 40, 40]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
    
    # Figer première ligne
    ws.freeze_panes = "A2"
    
    return wb, ws

def remplir_ligne_centrale(ws, row_idx, centrale, data_audit=None):
    """Remplit une ligne avec données centrale et audit"""
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Données de base (depuis ANNEXE 1)
    data = [
        centrale.get('id', ''),
        centrale.get('nom', ''),
        centrale.get('puissance_kwc', ''),
        centrale.get('type', ''),
        centrale.get('dept', ''),
        datetime.now().strftime("%Y-%m-%d"),  # Date audit (à compléter)
        "DiagPV",  # Auditeur
        "À compléter",  # Statut (sera rempli depuis checklist JSON)
        0,  # Anomalies critiques (sera calculé depuis checklist)
        0,  # Anomalies majeures
        0,  # Anomalies mineures
        0,  # Photos prises
        "À vérifier",  # Conformité IEC
        "À compléter",  # Recommandations
        "À compléter"   # Observations
    ]
    
    # Si data_audit fournie (depuis JSON checklist), compléter
    if data_audit:
        data[7] = data_audit.get('statut', 'À compléter')
        data[8] = data_audit.get('anomalies_critiques', 0)
        data[9] = data_audit.get('anomalies_majeures', 0)
        data[10] = data_audit.get('anomalies_mineures', 0)
        data[11] = data_audit.get('nb_photos', 0)
        data[12] = data_audit.get('conformite_iec', 'À vérifier')
        data[13] = data_audit.get('recommandations', 'À compléter')
        data[14] = data_audit.get('observations', 'À compléter')
    
    # Écrire ligne
    for col_idx, value in enumerate(data, 1):
        cell = ws.cell(row_idx, col_idx, value)
        cell.border = border
        cell.alignment = Alignment(vertical='top', wrap_text=True)
        
        # Alignement spécifique colonnes numériques
        if col_idx in [3, 9, 10, 11, 12]:
            cell.alignment = Alignment(horizontal='center', vertical='center')

def main():
    """Génère ANNEXE 2 automatiquement"""
    
    print("📊 GÉNÉRATEUR ANNEXE 2 AUTOMATIQUE - GIRASOLE 2025\n")
    
    # 1. Charger liste 52 centrales
    base_dir = Path('/home/user/girasole_mission_2025')
    centrales_path = base_dir / 'data' / 'liste_52_centrales.json'
    print(f"📄 Lecture liste centrales: {centrales_path.name}")
    with open(centrales_path, 'r', encoding='utf-8') as f:
        centrales = json.load(f)
    print(f"   ✅ {len(centrales)} centrales chargées\n")
    
    # 2. Créer structure ANNEXE 2
    print("📋 Création structure ANNEXE 2...")
    wb, ws = creer_annexe2_structure()
    print("   ✅ Structure créée\n")
    
    # 3. Remplir lignes pour chaque centrale
    print("⚙️  Remplissage données centrales...\n")
    
    json_exports_dir = base_dir / 'exports_json'
    
    # Créer répertoire exports si n'existe pas
    json_exports_dir.mkdir(exist_ok=True)
    
    centrales_avec_data = 0
    centrales_sans_data = 0
    
    for i, centrale in enumerate(centrales, 1):
        row_idx = i + 1
        centrale_id = centrale['id']
        
        # Chercher JSON terrain et BE
        json_terrain_path = json_exports_dir / f"{centrale_id}_terrain.json"
        json_be_path = json_exports_dir / f"{centrale_id}_be.json"
        
        # Charger data audit si disponible
        data_audit = None
        if json_terrain_path.exists() or json_be_path.exists():
            data_audit = charger_data_audit_json(json_terrain_path, json_be_path)
            status_icon = "✅📊"
            centrales_avec_data += 1
        else:
            status_icon = "⚪"
            centrales_sans_data += 1
        
        remplir_ligne_centrale(ws, row_idx, centrale, data_audit=data_audit)
        
        print(f"   {i:2d}. {status_icon} {centrale['id']} - {centrale['nom']}")
    
    # 4. Statistiques
    print(f"\n📊 STATISTIQUES:")
    print(f"   ✅ {centrales_avec_data} centrales avec données audit")
    print(f"   ⚪ {centrales_sans_data} centrales sans données (à compléter)")
    
    # 5. Sauvegarder
    output_dir = base_dir / 'outputs_annexe2'
    output_dir.mkdir(exist_ok=True)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = output_dir / f"ANNEXE2_GIRASOLE_SYNTHESE_{timestamp}.xlsx"
    wb.save(output_path)
    
    print(f"\n✅ ANNEXE 2 GÉNÉRÉE AVEC SUCCÈS!")
    print(f"📁 Fichier: {output_path}")
    print(f"📊 52 centrales incluses")
    print(f"✅ Format conforme exemple Girasole")
    print(f"\n💡 MODE D'EMPLOI:")
    print(f"   1. Les auditeurs exportent checklists terrain en JSON")
    print(f"   2. Placer les JSON dans: {json_exports_dir}")
    print(f"   3. Relancer ce script pour mise à jour automatique")
    print(f"   4. Format attendu: [ID_CENTRALE]_terrain.json et [ID_CENTRALE]_be.json")

if __name__ == "__main__":
    main()
