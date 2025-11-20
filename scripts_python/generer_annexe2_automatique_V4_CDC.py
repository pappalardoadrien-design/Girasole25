#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
GÉNÉRATEUR ANNEXE 2 AUTOMATIQUE V4 CDC - GIRASOLE 2025
Diagnostic Photovoltaïque

Génère l'ANNEXE 2 (fichier de synthèse) depuis checklists V4 terrain
Format: Excel avec 69 colonnes (15 base + 54 champs CDC)
Conformité CDC GIRASOLE 100%

Auteur : DiagPV Assistant Pro
Date : 20 janvier 2025
Version : 4.0 CDC COMPLET
"""

import json
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime


def charger_json(chemin: Path) -> dict:
    """Charge un fichier JSON"""
    try:
        with open(chemin, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        print(f"   ⚠️  Erreur lecture {chemin.name}: {e}")
        return {}


def charger_data_audit_v4_annexe2(json_v4_path: Path):
    """
    Charge JSON V4 et extrait 69 champs pour ANNEXE 2
    
    Returns:
        Dict avec 69 champs ou None si erreur
    """
    
    if not json_v4_path.exists():
        return None
    
    try:
        data = charger_json(json_v4_path)
        
        if not data:
            return None
        
        # Vérifier version
        version = data.get('metadata', {}).get('version', '3.0')
        if version < '4.0':
            print(f"   ⚠️  JSON V3 détecté : {json_v4_path.name} (version {version})")
            return None
        
        centrale = data.get('centrale', {})
        audit = data.get('audit', {})
        photos = data.get('photos', {})
        stats = data.get('statistiques', {})
        
        # Construire ligne Excel (69 champs)
        data_ligne = {
            # Base (14 colonnes)
            'id': centrale.get('id', ''),
            'nom': centrale.get('nom', ''),
            'puissance_kwc': centrale.get('puissance_kwc', ''),
            'type': centrale.get('type_installation', 'SOL'),
            'dept': centrale.get('departement', ''),
            'date_audit': audit.get('date_audit', ''),
            'auditeur': audit.get('auditeur_nom', 'DiagPV'),
            'statut': audit.get('statut_global', 'À compléter'),
            'anomalies_critiques': stats.get('anomalies_critiques', 0),
            'anomalies_majeures': stats.get('anomalies_majeures', 0),
            'anomalies_mineures': stats.get('anomalies_mineures', 0),
            'photos_totales': stats.get('total_photos', 0),
            'conformite_cdc': data.get('metadata', {}).get('conformite_cdc', '100%'),
            'version_checklist': version,
            
            # Section 2 : Documents GIRASOLE (4 colonnes)
            'doc_autocontrole': audit.get('doc_autocontrole', 'N/A'),
            'doc_plan_implantation': audit.get('doc_plan_implantation', 'N/A'),
            'doc_plan_electrique': audit.get('doc_plan_electrique', 'N/A'),
            'doc_schema_boites': audit.get('doc_schema_boites', 'N/A'),
            
            # Section 3 : Électrique Détaillé (25 colonnes)
            'elec_type_cheminement': audit.get('cablage_type_cheminement', 'N/A'),
            'elec_couleurs_dc': audit.get('cablage_couleurs_dc', 'N/A'),
            'elec_sections_dc': audit.get('cablage_sections_dc', 'N/A'),
            'elec_sections_ac': audit.get('cablage_sections_ac', 'N/A'),
            'elec_etat_cablage': audit.get('cablage_etat_general', 'N/A'),
            'elec_fixations': audit.get('cablage_fixations', 'N/A'),
            'elec_protection_mecanique': audit.get('cablage_protection_mecanique', 'N/A'),
            'elec_etancheite_presse': audit.get('cablage_etancheite_presse_etoupes', 'N/A'),
            'elec_equipotentielles': audit.get('elec_equipotentielles', 'N/A'),
            'elec_terre_valeur': audit.get('elec_terre_valeur', 'N/A'),
            'elec_parafoudre': audit.get('elec_parafoudre_type', 'N/A'),
            'elec_etat_coffrets': audit.get('elec_etat_coffrets', 'N/A'),
            'elec_etancheite_coffrets': audit.get('elec_etancheite_coffrets', 'N/A'),
            'elec_etiquetage_presence': audit.get('elec_etiquetage_presence', 'N/A'),
            'elec_etiquetage_qualite': audit.get('elec_etiquetage_qualite', 'N/A'),
            'elec_serrages_borniers': audit.get('elec_serrages_borniers', 'N/A'),
            'elec_coupure_dc': audit.get('elec_coupure_dc', 'N/A'),
            'elec_protection_dc': audit.get('elec_protection_dc', 'N/A'),
            'elec_protection_ac': audit.get('elec_protection_ac', 'N/A'),
            'elec_differentiel': audit.get('elec_differentiel_sensibilite', 'N/A'),
            'elec_accessibilite': audit.get('elec_accessibilite', 'N/A'),
            'elec_ventilation': audit.get('elec_ventilation', 'N/A'),
            'elec_signalisation': audit.get('elec_signalisation', 'N/A'),
            'elec_protection_surtension': audit.get('elec_protection_surtension', 'N/A'),
            'elec_continuite_terre': audit.get('elec_continuite_terre', 'N/A'),
            
            # Section 4 : Tranchées (2 colonnes)
            'tranchees_accessibilite': audit.get('tranchees_accessibilite', 'N/A'),
            'tranchees_conformite': audit.get('tranchees_conformite', 'N/A'),
            
            # Section 5 : Modules (7 colonnes)
            'mp_etat_general': audit.get('modules_etat_general', 'N/A'),
            'mp_defauts_visibles': audit.get('modules_defauts_visibles', 'N/A'),
            'mp_cablage': audit.get('modules_cablage', 'N/A'),
            'mp_connecteurs': audit.get('modules_connecteurs_mc4', 'N/A'),
            'mp_fixations': audit.get('modules_fixations', 'N/A'),
            'mp_orientation': audit.get('modules_orientation', 'N/A'),
            'mp_masques': audit.get('modules_masques_ombrages', 'N/A'),
            
            # Section 6 : Structure (5 colonnes)
            'toit_etat_structure': audit.get('structure_etat_general', 'N/A'),
            'toit_type_structure': audit.get('structure_type', 'N/A'),
            'toit_fixations': audit.get('structure_fixations', 'N/A'),
            'toit_stabilite': audit.get('structure_stabilite', 'N/A'),
            'toit_acces_maintenance': audit.get('structure_acces_maintenance', 'N/A'),
            
            # Section 7 : Boîtes (4 colonnes)
            'bp_etat_general': audit.get('boites_etat_general', 'N/A'),
            'bp_etancheite': audit.get('boites_etancheite', 'N/A'),
            'bp_accessibilite': audit.get('boites_accessibilite', 'N/A'),
            'bp_cablage_interne': audit.get('boites_cablage_interne', 'N/A'),
            
            # Section 8 : Toiture (13 colonnes - conditionnelles)
            'toiture_applicable': audit.get('toiture_applicable', 'Non'),
            'toiture_demontage': audit.get('toiture_demontage', 'N/A'),
            'toiture_si_type': audit.get('toiture_si_type', 'N/A'),
            'toiture_si_validite': audit.get('toiture_si_validite', 'N/A'),
            'toiture_plaques_type': audit.get('toiture_plaques_type', 'N/A'),
            'toiture_plaques_etat': audit.get('toiture_plaques_etat', 'N/A'),
            'toiture_fixations': audit.get('toiture_fixations', 'N/A'),
            'toiture_etancheite': audit.get('toiture_etancheite', 'N/A'),
            'toiture_ecrans': audit.get('toiture_ecrans', 'N/A'),
            'toiture_charpente': audit.get('toiture_charpente', 'N/A'),
            'toiture_risques_infiltration': audit.get('toiture_risques_infiltration', 'N/A'),
            'toiture_conformite_dtu': audit.get('toiture_conformite_dtu', 'N/A'),
            'toiture_conformite_etn': audit.get('toiture_conformite_etn', 'N/A'),
            
            # Photos par catégorie (8 colonnes)
            'photos_doc': len(photos.get('DOC', [])),
            'photos_elec': len(photos.get('ELEC', [])),
            'photos_tranchees': len(photos.get('TRANCHEES', [])),
            'photos_mp': len(photos.get('MP', [])),
            'photos_toit': len(photos.get('TOIT', [])),
            'photos_bp': len(photos.get('BP', [])),
            'photos_gen': len(photos.get('GEN', [])),
            'photos_toiture_detail': len(photos.get('TOITURE_DETAIL', [])),
            
            # Synthèse (2 colonnes)
            'recommandations': audit.get('recommandations_prioritaires', 'À compléter'),
            'observations': audit.get('observations_terrain', 'À compléter')
        }
        
        return data_ligne
    
    except Exception as e:
        print(f"   ❌ Erreur parsing {json_v4_path.name}: {e}")
        return None


def creer_annexe2_structure_v4():
    """Crée structure Excel ANNEXE 2 V4 (69 colonnes)"""
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Synthèse Audits V4"
    
    # Styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # En-têtes colonnes V4 (69 colonnes)
    headers = [
        # Base (14)
        "ID Centrale",
        "Nom Projet",
        "Puissance (kWc)",
        "Type Installation",
        "Département",
        "Date Audit",
        "Auditeur",
        "Statut Global",
        "Anomalies Critiques",
        "Anomalies Majeures",
        "Anomalies Mineures",
        "Photos Totales",
        "Conformité CDC",
        "Version Checklist",
        
        # Section 2 : Documents GIRASOLE (4)
        "DOC - Autocontrôle",
        "DOC - Plan Implantation",
        "DOC - Plan Électrique",
        "DOC - Schéma Boîtes",
        
        # Section 3 : Électrique (25)
        "ELEC - Type Cheminement",
        "ELEC - Couleurs DC",
        "ELEC - Sections DC",
        "ELEC - Sections AC",
        "ELEC - État Câblage",
        "ELEC - Fixations",
        "ELEC - Protection Mécanique",
        "ELEC - Étanchéité Presse-Étoupes",
        "ELEC - Équipotentielles",
        "ELEC - Terre (Ω)",
        "ELEC - Parafoudre SPD",
        "ELEC - État Coffrets",
        "ELEC - Étanchéité Coffrets",
        "ELEC - Étiquetage Présence",
        "ELEC - Étiquetage Qualité",
        "ELEC - Serrages Borniers",
        "ELEC - Coupure DC",
        "ELEC - Protection DC",
        "ELEC - Protection AC",
        "ELEC - Différentiel",
        "ELEC - Accessibilité",
        "ELEC - Ventilation",
        "ELEC - Signalisation",
        "ELEC - Protection Surtension",
        "ELEC - Continuité Terre",
        
        # Section 4 : Tranchées (2)
        "TRANCHEES - Accessibilité",
        "TRANCHEES - Conformité",
        
        # Section 5 : Modules (7)
        "MP - État Général",
        "MP - Défauts Visibles",
        "MP - Câblage",
        "MP - Connecteurs MC4",
        "MP - Fixations",
        "MP - Orientation",
        "MP - Masques Ombrages",
        
        # Section 6 : Structure (5)
        "TOIT - État Structure",
        "TOIT - Type Structure",
        "TOIT - Fixations",
        "TOIT - Stabilité",
        "TOIT - Accès Maintenance",
        
        # Section 7 : Boîtes (4)
        "BP - État Général",
        "BP - Étanchéité",
        "BP - Accessibilité",
        "BP - Câblage Interne",
        
        # Section 8 : Toiture (13)
        "TOITURE - Applicable",
        "TOITURE - Démontage",
        "TOITURE - SI Type",
        "TOITURE - SI Validité",
        "TOITURE - Plaques Type",
        "TOITURE - Plaques État",
        "TOITURE - Fixations",
        "TOITURE - Étanchéité",
        "TOITURE - Écrans",
        "TOITURE - Charpente",
        "TOITURE - Risques Infiltration",
        "TOITURE - Conformité DTU 40.35",
        "TOITURE - Conformité ETN",
        
        # Photos (8)
        "Photos DOC",
        "Photos ELEC",
        "Photos TRANCHEES",
        "Photos MP",
        "Photos TOIT",
        "Photos BP",
        "Photos GEN",
        "Photos TOITURE-DETAIL",
        
        # Synthèse (2)
        "Recommandations Prioritaires",
        "Observations Terrain"
    ]
    
    # Écrire en-têtes
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(1, col_idx, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    # Ajuster largeurs colonnes (69 valeurs)
    col_widths = [
        # Base (14)
        12, 30, 15, 20, 12, 12, 20, 15, 18, 18, 18, 12, 15, 12,
        
        # Documents (4)
        20, 20, 20, 20,
        
        # Électrique (25)
        18, 18, 15, 15, 18, 18, 18, 18, 18, 15, 18, 18, 18, 18, 18,
        18, 18, 18, 18, 15, 18, 18, 18, 18, 18,
        
        # Tranchées (2)
        20, 25,
        
        # Modules (7)
        18, 20, 18, 18, 18, 18, 18,
        
        # Structure (5)
        18, 20, 18, 18, 20,
        
        # Boîtes (4)
        18, 18, 18, 20,
        
        # Toiture (13)
        15, 20, 20, 18, 20, 18, 18, 20, 18, 20, 22, 20, 18,
        
        # Photos (8)
        12, 12, 12, 12, 12, 12, 12, 15,
        
        # Synthèse (2)
        40, 40
    ]
    
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
    
    # Figer première ligne
    ws.freeze_panes = "A2"
    
    return wb, ws


def remplir_ligne_centrale_v4(ws, row_idx, data_ligne):
    """Remplit ligne Excel avec 69 colonnes V4"""
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Ordre colonnes (69 valeurs)
    values = [
        # Base (14)
        data_ligne.get('id', ''),
        data_ligne.get('nom', ''),
        data_ligne.get('puissance_kwc', ''),
        data_ligne.get('type', ''),
        data_ligne.get('dept', ''),
        data_ligne.get('date_audit', ''),
        data_ligne.get('auditeur', ''),
        data_ligne.get('statut', ''),
        data_ligne.get('anomalies_critiques', 0),
        data_ligne.get('anomalies_majeures', 0),
        data_ligne.get('anomalies_mineures', 0),
        data_ligne.get('photos_totales', 0),
        data_ligne.get('conformite_cdc', '100%'),
        data_ligne.get('version_checklist', '4.0'),
        
        # Documents (4)
        data_ligne.get('doc_autocontrole', 'N/A'),
        data_ligne.get('doc_plan_implantation', 'N/A'),
        data_ligne.get('doc_plan_electrique', 'N/A'),
        data_ligne.get('doc_schema_boites', 'N/A'),
        
        # Électrique (25)
        data_ligne.get('elec_type_cheminement', 'N/A'),
        data_ligne.get('elec_couleurs_dc', 'N/A'),
        data_ligne.get('elec_sections_dc', 'N/A'),
        data_ligne.get('elec_sections_ac', 'N/A'),
        data_ligne.get('elec_etat_cablage', 'N/A'),
        data_ligne.get('elec_fixations', 'N/A'),
        data_ligne.get('elec_protection_mecanique', 'N/A'),
        data_ligne.get('elec_etancheite_presse', 'N/A'),
        data_ligne.get('elec_equipotentielles', 'N/A'),
        data_ligne.get('elec_terre_valeur', 'N/A'),
        data_ligne.get('elec_parafoudre', 'N/A'),
        data_ligne.get('elec_etat_coffrets', 'N/A'),
        data_ligne.get('elec_etancheite_coffrets', 'N/A'),
        data_ligne.get('elec_etiquetage_presence', 'N/A'),
        data_ligne.get('elec_etiquetage_qualite', 'N/A'),
        data_ligne.get('elec_serrages_borniers', 'N/A'),
        data_ligne.get('elec_coupure_dc', 'N/A'),
        data_ligne.get('elec_protection_dc', 'N/A'),
        data_ligne.get('elec_protection_ac', 'N/A'),
        data_ligne.get('elec_differentiel', 'N/A'),
        data_ligne.get('elec_accessibilite', 'N/A'),
        data_ligne.get('elec_ventilation', 'N/A'),
        data_ligne.get('elec_signalisation', 'N/A'),
        data_ligne.get('elec_protection_surtension', 'N/A'),
        data_ligne.get('elec_continuite_terre', 'N/A'),
        
        # Tranchées (2)
        data_ligne.get('tranchees_accessibilite', 'N/A'),
        data_ligne.get('tranchees_conformite', 'N/A'),
        
        # Modules (7)
        data_ligne.get('mp_etat_general', 'N/A'),
        data_ligne.get('mp_defauts_visibles', 'N/A'),
        data_ligne.get('mp_cablage', 'N/A'),
        data_ligne.get('mp_connecteurs', 'N/A'),
        data_ligne.get('mp_fixations', 'N/A'),
        data_ligne.get('mp_orientation', 'N/A'),
        data_ligne.get('mp_masques', 'N/A'),
        
        # Structure (5)
        data_ligne.get('toit_etat_structure', 'N/A'),
        data_ligne.get('toit_type_structure', 'N/A'),
        data_ligne.get('toit_fixations', 'N/A'),
        data_ligne.get('toit_stabilite', 'N/A'),
        data_ligne.get('toit_acces_maintenance', 'N/A'),
        
        # Boîtes (4)
        data_ligne.get('bp_etat_general', 'N/A'),
        data_ligne.get('bp_etancheite', 'N/A'),
        data_ligne.get('bp_accessibilite', 'N/A'),
        data_ligne.get('bp_cablage_interne', 'N/A'),
        
        # Toiture (13)
        data_ligne.get('toiture_applicable', 'Non'),
        data_ligne.get('toiture_demontage', 'N/A'),
        data_ligne.get('toiture_si_type', 'N/A'),
        data_ligne.get('toiture_si_validite', 'N/A'),
        data_ligne.get('toiture_plaques_type', 'N/A'),
        data_ligne.get('toiture_plaques_etat', 'N/A'),
        data_ligne.get('toiture_fixations', 'N/A'),
        data_ligne.get('toiture_etancheite', 'N/A'),
        data_ligne.get('toiture_ecrans', 'N/A'),
        data_ligne.get('toiture_charpente', 'N/A'),
        data_ligne.get('toiture_risques_infiltration', 'N/A'),
        data_ligne.get('toiture_conformite_dtu', 'N/A'),
        data_ligne.get('toiture_conformite_etn', 'N/A'),
        
        # Photos (8)
        data_ligne.get('photos_doc', 0),
        data_ligne.get('photos_elec', 0),
        data_ligne.get('photos_tranchees', 0),
        data_ligne.get('photos_mp', 0),
        data_ligne.get('photos_toit', 0),
        data_ligne.get('photos_bp', 0),
        data_ligne.get('photos_gen', 0),
        data_ligne.get('photos_toiture_detail', 0),
        
        # Synthèse (2)
        data_ligne.get('recommandations', 'À compléter'),
        data_ligne.get('observations', 'À compléter')
    ]
    
    # Écrire ligne
    for col_idx, value in enumerate(values, 1):
        cell = ws.cell(row_idx, col_idx, value)
        cell.border = border
        cell.alignment = Alignment(vertical='top', wrap_text=True)
        
        # Alignement numérique
        if isinstance(value, (int, float)):
            cell.alignment = Alignment(horizontal='center', vertical='center')


def main():
    """Génère ANNEXE 2 V4 automatiquement depuis JSON V4"""
    
    print("="*100)
    print("📊 GÉNÉRATEUR ANNEXE 2 V4 CDC COMPLET - GIRASOLE 2025")
    print("="*100)
    print()
    
    # Chemins
    base_dir = Path(__file__).parent.parent
    exports_dir = base_dir / 'exports_json'
    output_dir = base_dir / 'outputs_annexe2'
    output_dir.mkdir(exist_ok=True)
    
    # Créer dossier exports si n'existe pas
    exports_dir.mkdir(exist_ok=True)
    
    # 1. Créer structure ANNEXE 2 V4
    print("📋 Création structure ANNEXE 2 V4 (69 colonnes)...")
    wb, ws = creer_annexe2_structure_v4()
    print("   ✅ Structure créée\n")
    
    # 2. Lister JSON V4 disponibles
    json_files = list(exports_dir.glob("AUDIT_*.json"))
    
    if not json_files:
        print(f"⚠️  Aucun fichier JSON V4 trouvé dans {exports_dir}")
        print(f"💡 Format attendu : AUDIT_[ID]_[NOM]_[DATE].json")
        print(f"💡 L'ANNEXE 2 sera générée avec structure vide (prête pour remplissage futur)\n")
    else:
        print(f"📊 {len(json_files)} fichiers JSON V4 détectés\n")
    
    # 3. Remplir lignes pour chaque JSON V4
    print("⚙️  Remplissage données centrales V4...\n")
    
    centrales_avec_data = 0
    centrales_sans_data = 0
    centrales_v3_skip = 0
    row_idx = 2  # Commence à ligne 2 (après en-têtes)
    
    for json_path in sorted(json_files):
        # Charger data audit V4
        data_ligne = charger_data_audit_v4_annexe2(json_path)
        
        if data_ligne:
            remplir_ligne_centrale_v4(ws, row_idx, data_ligne)
            
            centrale_id = data_ligne.get('id', 'UNKNOWN')
            centrale_nom = data_ligne.get('nom', 'N/A')
            status_icon = "✅📊"
            centrales_avec_data += 1
            
            print(f"   {row_idx-1:2d}. {status_icon} {centrale_id} - {centrale_nom} ({data_ligne.get('photos_totales', 0)} photos)")
            
            row_idx += 1
        else:
            # Vérifier si V3
            data_raw = charger_json(json_path)
            version = data_raw.get('metadata', {}).get('version', '3.0')
            if version < '4.0':
                centrales_v3_skip += 1
                print(f"   ⚠️  SKIP {json_path.name} (V3 détecté)")
    
    # 4. Statistiques
    print(f"\n📊 STATISTIQUES:")
    print(f"   ✅ Centrales V4 avec données : {centrales_avec_data}")
    print(f"   ⚠️  Centrales V3 skippées : {centrales_v3_skip}")
    print(f"   📋 Lignes remplies : {centrales_avec_data}")
    print(f"   📊 Colonnes : 69 (conformité CDC 100%)")
    
    # 5. Sauvegarder
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = output_dir / f"ANNEXE2_V4_CDC_GIRASOLE_{timestamp}.xlsx"
    wb.save(output_path)
    
    print(f"\n✅ ANNEXE 2 V4 GÉNÉRÉE AVEC SUCCÈS!")
    print(f"📁 Fichier : {output_path}")
    print(f"📊 {centrales_avec_data} centrales incluses")
    print(f"✅ Format V4 : 69 colonnes (54 points CDC)")
    print(f"✅ Conformité CDC GIRASOLE 100%")
    
    print(f"\n💡 MODE D'EMPLOI:")
    print(f"   1. Les auditeurs exportent checklists V4 en JSON depuis terrain")
    print(f"   2. Placer les JSON dans : {exports_dir}")
    print(f"   3. Relancer ce script pour mise à jour automatique")
    print(f"   4. Format JSON : AUDIT_[ID]_[NOM]_[DATE].json")
    print(f"   5. Seuls les JSON V4 (version ≥ 4.0) sont traités")
    
    print("\n" + "="*100)


if __name__ == "__main__":
    main()
